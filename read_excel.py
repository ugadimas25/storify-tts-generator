import openpyxl
wb = openpyxl.load_workbook('books_list.xlsx')
ws = wb.active
headers = [cell.value for cell in ws[1]]
print('Headers:', headers)
print()
for row in ws.iter_rows(min_row=2, values_only=True):
    print(list(row))
