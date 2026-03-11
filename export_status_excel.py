"""
Export status semua buku (PDF phase1, PDF phase2, Gutenberg) ke Excel.

Kolom: No | Book ID | Source | Status | Chapters | Summaries | Audio

Usage:
    python export_status_excel.py
    python export_status_excel.py --output my_status.xlsx
"""
import argparse
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    raise

BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"

# Fill colors per status
COLOR_DONE    = "C6EFCE"  # green
COLOR_PARTIAL = "FFEB9C"  # yellow
COLOR_PENDING = "FFC7CE"  # red


def check_book(book_name: str, source: str) -> dict:
    done_marker   = DATA_DIR / "audio"     / book_name / ".done"
    chapters_dir  = DATA_DIR / "chapters"  / book_name
    summaries_dir = DATA_DIR / "summaries" / book_name
    audio_dir     = DATA_DIR / "audio"     / book_name

    n_ch  = len(list(chapters_dir.glob("chapter_*.txt")))   if chapters_dir.exists()  else 0
    n_sum = len(list(summaries_dir.glob("chapter_*.json"))) if summaries_dir.exists() else 0
    n_aud = len(list(audio_dir.glob("chapter_*.mp3")))      if audio_dir.exists()     else 0

    if done_marker.exists():
        status = "DONE"
    elif n_ch > 0 and n_ch == n_sum == n_aud:
        status = "DONE"
    elif n_sum > 0 or n_aud > 0:
        status = "PARTIAL"
    else:
        status = "PENDING"

    return {
        "book_id": book_name,
        "source": source,
        "status": status,
        "chapters": n_ch,
        "summaries": n_sum,
        "audio": n_aud,
    }


def collect_all() -> list[dict]:
    rows = []

    # PDF phase1
    books1 = BASE_DIR / "src" / "books"
    if books1.exists():
        for pdf in sorted(books1.glob("*.pdf")):
            rows.append(check_book(pdf.stem, "PDF-phase1"))

    # PDF phase2
    books2 = BASE_DIR / "src" / "books_phase2"
    if books2.exists():
        for pdf in sorted(books2.glob("*.pdf")):
            rows.append(check_book(pdf.stem, "PDF-phase2"))

    # Gutenberg
    gutenberg = BASE_DIR / "gutenberg_books" / "plain_text"
    if gutenberg.exists():
        for txt in sorted(gutenberg.glob("*.txt")):
            if txt.name == "index.txt":
                continue
            book_name = txt.stem.replace("_", "")
            rows.append(check_book(book_name, "Gutenberg"))

    return rows


def export_excel(rows: list[dict], output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Status"

    # Header
    headers = ["No", "Book ID", "Source", "Status", "Chapters", "Summaries", "Audio"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 20

    # Data rows
    status_color = {
        "DONE":    COLOR_DONE,
        "PARTIAL": COLOR_PARTIAL,
        "PENDING": COLOR_PENDING,
    }

    for i, row in enumerate(rows, 1):
        r = i + 1
        color = status_color.get(row["status"], "FFFFFF")
        fill = PatternFill("solid", fgColor=color)

        values = [
            i,
            row["book_id"],
            row["source"],
            row["status"],
            row["chapters"] or "",
            row["summaries"] or "",
            row["audio"] or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center" if col != 2 else "left")

    # Summary rows
    summary_row = len(rows) + 3
    total   = len(rows)
    done    = sum(1 for r in rows if r["status"] == "DONE")
    partial = sum(1 for r in rows if r["status"] == "PARTIAL")
    pending = sum(1 for r in rows if r["status"] == "PENDING")

    ws.cell(row=summary_row,     column=1, value="Total").font = Font(bold=True)
    ws.cell(row=summary_row,     column=2, value=total)
    ws.cell(row=summary_row + 1, column=1, value="DONE").font  = Font(bold=True, color="375623")
    ws.cell(row=summary_row + 1, column=2, value=done)
    ws.cell(row=summary_row + 1, column=1).fill = PatternFill("solid", fgColor=COLOR_DONE)
    ws.cell(row=summary_row + 2, column=1, value="PARTIAL").font = Font(bold=True, color="7D6608")
    ws.cell(row=summary_row + 2, column=2, value=partial)
    ws.cell(row=summary_row + 2, column=1).fill = PatternFill("solid", fgColor=COLOR_PARTIAL)
    ws.cell(row=summary_row + 3, column=1, value="PENDING").font = Font(bold=True, color="9C0006")
    ws.cell(row=summary_row + 3, column=2, value=pending)
    ws.cell(row=summary_row + 3, column=1).fill = PatternFill("solid", fgColor=COLOR_PENDING)

    # Column widths
    col_widths = [6, 16, 14, 10, 10, 12, 8]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"Saved: {output_path}")
    print(f"Total: {total}  |  DONE: {done}  |  PARTIAL: {partial}  |  PENDING: {pending}")


def main():
    parser = argparse.ArgumentParser(description="Export book processing status to Excel")
    parser.add_argument("--output", type=str, default=None, help="Output Excel file path")
    args = parser.parse_args()

    output_path = Path(args.output) if args.output else BASE_DIR / f"status_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    rows = collect_all()
    export_excel(rows, output_path)


if __name__ == "__main__":
    main()
