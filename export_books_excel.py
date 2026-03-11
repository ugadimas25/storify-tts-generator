"""
Export list of all PDF books to Excel with OCR classification columns.

Usage:
    python export_books_excel.py                                    # src/books -> books_list.xlsx
    python export_books_excel.py --books-dir src/books_phase2       # phase2 -> books_list_phase2.xlsx
    python export_books_excel.py --books-dir src/books_phase2 --output my_list.xlsx

Kolom yang dihasilkan:
    No | File Name | Book ID | Size | Last Modified | OCR | just cover
    - Isi kolom OCR dan just cover secara manual di Excel:
        OCR = "V"         -> seluruh PDF di-OCR (PDF scan)
        just cover = "V"  -> hanya cover yang image, konten bisa di-extract biasa
        keduanya kosong   -> PDF teks murni, tidak perlu OCR
"""
import argparse
import sys
from pathlib import Path
from datetime import datetime

sys.path.insert(0, ".")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter


def get_pdf_size(path: Path) -> str:
    size_bytes = path.stat().st_size
    if size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    return f"{size_bytes / (1024 * 1024):.1f} MB"


def main():
    parser = argparse.ArgumentParser(description="Export PDF book list to Excel")
    parser.add_argument(
        "--books-dir", type=str, default="src/books",
        help="Directory containing PDF files (default: src/books)"
    )
    parser.add_argument(
        "--output", type=str, default=None,
        help="Output Excel filename (default: books_list.xlsx, or books_list_<dirname>.xlsx for non-default dir)"
    )
    args = parser.parse_args()

    books_dir = Path(args.books_dir)

    # Auto-generate output filename based on books dir
    if args.output:
        output_file = Path(args.output)
    elif str(books_dir) == "src/books":
        output_file = Path("books_list.xlsx")
    else:
        safe_name = books_dir.name.replace(" ", "_")
        output_file = Path(f"books_list_{safe_name}.xlsx")

    pdf_files = sorted(books_dir.glob("*.pdf"))
    if not pdf_files:
        print(f"No PDF files found in {books_dir}")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Books"

    # Header style
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    headers    = ["No", "File Name", "Book ID", "Size", "Last Modified", "OCR", "just cover"]
    col_widths = [6,    25,          12,        12,     22,              10,    12]

    for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 25

    # OCR / just cover header hint colors
    ocr_header_fill      = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
    cover_header_fill    = PatternFill(start_color="7B3F00", end_color="7B3F00", fill_type="solid")
    ws.cell(row=1, column=6).fill = ocr_header_fill
    ws.cell(row=1, column=7).fill = cover_header_fill

    # Alternating row colors
    fill_even = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    fill_odd  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for i, pdf_path in enumerate(pdf_files, start=1):
        row  = i + 1
        fill = fill_even if i % 2 == 0 else fill_odd
        modified = datetime.fromtimestamp(pdf_path.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
        book_id  = pdf_path.stem

        data = [i, pdf_path.name, book_id, get_pdf_size(pdf_path), modified, None, None]
        for col, value in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", horizontal="center" if col in (1, 3, 4, 6, 7) else "left")

    # Auto filter & freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"

    wb.save(output_file)
    print(f"Saved: {output_file} ({len(pdf_files)} files)")
    print(f"  -> Isi kolom 'OCR' dengan 'V' untuk PDF scan penuh")
    print(f"  -> Isi kolom 'just cover' dengan 'V' jika hanya cover yang image")
    print(f"  -> Biarkan kosong jika PDF teks murni")


if __name__ == "__main__":
    main()
